#! python3
# text_file_to_spreadsheet.py - Reads in the contents of several text files, and inserts those
#                               contents into a spreadsheet, with one line of text per row.
#                               The lines of the first text file will be in column A, second
#                               file in column B, etc...
#                               Note: text files must be stored in "text_files" folder, located in
#                               the same folder as text_file_to_spreadsheet.py
# Usage: text_file_to_spreadsheet.py [**text_file_names]        - specifically name each file to use, or
#        text_file_to_spreadsheet.py all                        - read all text files in text_files folder

import sys
import os
import openpyxl
from openpyxl.styles import Font

# Validate command-line usage
try:
    if len(sys.argv) < 2:
        raise Exception

    # Validate command-line arguments
    if sys.argv[1] == "all":
        all_files = os.listdir("text_files")
    else:
        all_files = sys.argv[1:]

    # Read file lines into dictionary
    file_contents = {}
    for file_name in all_files:
        text = open(os.path.join("text_files", file_name))
        file_contents[file_name] = text.readlines()

    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Text Contents"

except Exception as e:
    print(
        f"\n{e}\nUsage: text_file_to_spreadsheet.py [**text_file_names], or\ntext_file_to_spreadsheet.py all\n")

row_num = 1
col_num = 1
for file in file_contents:
    row_num = 1
    sheet.cell(row=row_num, column=col_num).value = file
    sheet.cell(row=row_num, column=col_num).font = Font(bold=True)
    row_num += 1
    for line in file_contents[file]:
        sheet.cell(row=row_num, column=col_num).value = line
        row_num += 1
    col_num += 1

wb.save("text_files.xlsx")

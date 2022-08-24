#! python3
# cell_inverter.py - Inverts the rows and columns of a spreadsheet, and outputs the result to a new spreadsheet
# Usage: python3 cell_inverter.py [workbook_name]

import openpyxl, sys
from copy import copy

# Validate usage and setup workbook to operate on
try:
    if len(sys.argv) != 2:
        raise Exception
    
    wb = openpyxl.load_workbook(sys.argv[1])
    sheet_from = wb.active  # sheet to copy from
    sheet_inverse = wb.create_sheet(f"Inverted {sheet_from.title}",)    # sheet to copy to
    # Setup new sheet to be NxM dimension, for original worksheet of MxN dimension
    sheet_inverse.insert_rows(1, sheet_from.max_column-1)
    sheet_inverse.insert_cols(1, sheet_from.max_row-1)
except:
    print("Incorrect command-line arguments.\nUsage: python3 cell_inverter.py [workbook_name]\n")

# Iterate through each row and column
for x in range(1, sheet_from.max_row+1):
    for y in range(1, sheet_from.max_column+1):
        cell_from = sheet_from.cell(row=x,column=y)
        cell_to = sheet_inverse.cell(row=y,column=x)
        cell_to.value = cell_from.value     # copy value from one cell to another

        # Copy rest of cell formatting if necessary
        if cell_from.has_style:
            cell_to.font = copy(cell_from.font)
            cell_to.border = copy(cell_from.border)
            cell_to.fill = copy(cell_from.fill)
            cell_to.number_format = copy(cell_from.number_format)
            cell_to.protection = copy(cell_from.protection)
            cell_to.alignment = copy(cell_from.alignment)

# Save new workbook
wb.save(f"{sys.argv[1].split('.')[0]}_inverse.xlsx")
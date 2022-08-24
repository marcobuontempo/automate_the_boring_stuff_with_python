#! python3
# muliplication_table.py - takes a number N and creates an NxN multiplication table in an Excel spreadsheet
# Usage: python3 muliplication_table.py [table_size]

import sys, openpyxl
from openpyxl.styles import Font

# Ensure correct command-line usage
try:
    if len(sys.argv)!=2:
        raise Exception
    N = int(sys.argv[1])
except:
    print("Incorrect command-line arguments.\nUsage: python3 muliplication_table.py [table_size]")
    sys.exit()

# Create new workbook
wb = openpyxl.Workbook()
wb.active.title = "Multiplication Table"
sheet = wb["Multiplication Table"]

# Insert all multiplication values
for row_num in range(1,N+1):
    for col_num in range(1,N+1):
        sheet.cell(row=row_num,column=col_num).value = row_num * col_num

# Insert row headings
sheet.insert_rows(1)
for num in range(1,N+1):
    sheet.cell(row=1,column=num).value = num
    sheet.cell(row=1,column=num).font = Font(bold=True)
# Insert column headings
sheet.insert_cols(1)
for num in range(1,N+1):
    sheet.cell(row=num+1,column=1).value = num
    sheet.cell(row=num+1,column=1).font = Font(bold=True)

# Save workbook
wb.save(f"multiplication_table_{N}.xlsx")
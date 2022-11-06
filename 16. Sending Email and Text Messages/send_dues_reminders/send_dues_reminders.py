#! python3
# send_dues_reminder.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

# Get credentials for email from command line.
email_address = sys.argv[1]
email_password = sys.argv[2]

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")

last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# Check each member's payment status.
unpaid_members = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# Login to email account.
smtp_obj = smtplib.SMTP("smtp.gmail.com", 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login(email_address, email_password)

# Send out reminder emails.
for name, email in unpaid_members.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" % (latest_month,name,latest_month)
    print("Sending email to %s..." % email)
    send_email_status = smtp_obj.sendmail(email_address, email, body)

    if send_email_status != {}:
        print("There was a problem sending email to %s: %s" % (email, send_email_status))

smtp_obj.quit()